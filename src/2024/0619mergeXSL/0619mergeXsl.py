import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def merge_and_deduplicate(input_path, output_path, sheet_name, col1, col2, target_col):
    # 使用 pandas 读取 .xls 文件
    df = pd.read_excel(input_path, sheet_name=sheet_name)

    # # 合并和去重 - 不保持相对顺序
    # def merge_keywords(row):
    #     keywords1 = set(str(row[col1]).split('|')) if pd.notna(row[col1]) else set()
    #     keywords2 = set(str(row[col2]).split('|')) if pd.notna(row[col2]) else set()
    #     # 去重并保持 col1 中关键词的相对顺序
    #     unique_keywords = list(keywords1) + [kw for kw in keywords2 if kw not in keywords1]
    #     return '|'.join(unique_keywords)
    #
    # df[target_col] = df.apply(merge_keywords, axis=1)

    # 合并和去重 - 保持相对顺序
    def merge_keywords(row):
        keywords1 = str(row[col1]).split('|') if pd.notna(row[col1]) else []
        keywords2 = str(row[col2]).split('|') if pd.notna(row[col2]) else []

        # 去重并保持相对顺序
        seen = set()
        merged_keywords = []

        # 先添加 col1 中的关键词，保留顺序
        for kw in keywords1:
            kw = kw.strip()
            if kw and kw not in seen:
                merged_keywords.append(kw)
                seen.add(kw)

        # 再添加 col2 中的关键词，不重复
        for kw in keywords2:
            kw = kw.strip()
            if kw and kw not in seen:
                merged_keywords.append(kw)
                seen.add(kw)

        return '|'.join(merged_keywords)

    df[target_col] = df.apply(merge_keywords, axis=1)

    # 使用 openpyxl 保存到 .xlsx 文件
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 打开刚保存的文件以设置列宽和自动换行
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]

    # 设置列宽
    column_widths = {
        'C': 64,  # 领域关键词
        'D': 64,  # shf - 20个
        'E': 64,  # 合并后的关键词
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置自动换行
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(output_path)

# 使用示例
input_path = 'C:/Users/yuhuo/Documents/Document/6.17-22/6.19/input.xls'  # 修改为你的输入 Excel 文件路径
output_path = 'C:/Users/yuhuo/Documents/Document/6.17-22/6.19/output.xlsx'  # 修改为你的输出 Excel 文件路径
sheet_name = '分类领域整理'  # 修改为你的工作表名称
col1 = '领域关键词'
col2 = 'shf - 20个'
target_col = 'merged'

merge_and_deduplicate(input_path, output_path, sheet_name, col1, col2, target_col)
